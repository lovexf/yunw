import re
import socket
import paramiko
import pymysql
from pythonping import ping
import os
import datetime
import pandas as pd
from openpyxl import load_workbook
# pip install pythonping
import chardet
import time
from paramiko.client import SSHClient, AutoAddPolicy
from paramiko import AuthenticationException
from paramiko.ssh_exception import NoValidConnectionsError
from multiprocessing import Process

# pip install paramiko –i https://pypi.douban.com/simple/ --trusted-host pypi.douban.com

gl_id = 0
gl_name = 0
gl_ip = 0
gl_user = 0
gl_passwd = 0

gl_port1 = ''
gl_port2 = ''
gl_port3 = ''
gg_port_1 = ''
gg_port_2 = ''
gg_port_3 = ''

gl_f_id = 1
gl_f_i = 0

cur = 0
sql = ''

f_path_1 = ''
f_path_2 = ''
f_path_3 = ''

gl_cpu = ''
gl_nc = ''
gl_data = ''
gl_data1 = ''
gl_ls = ''
gl_null = ''

f_qy = ''

root_dir = '/'

target_date = datetime.date.today()
target_data = target_date.strftime("%Y%m%d")
target_date1 = datetime.date.today() - datetime.timedelta(days=1)
target_date1 = target_date1.strftime("%Y%m%d")

filename = '巡检%s' % target_date
fwq_name = ''
beif1 = ''
beif2 = ''
beif3 = ''


class SshClient():
    def __init__(self):
        self.ssh_client = SSHClient()

    # 连接服务器
    def ssh_login(self, host_ip, username, password):
        try:
            # 设置允许连接known_hosts文件中的主机（默认连接不在known_hosts文件中的主机会拒绝连接抛出SSHException）
            self.ssh_client.set_missing_host_key_policy(AutoAddPolicy())
            self.ssh_client.connect(host_ip, port=22, username=username, password=password, timeout=15)
        except AuthenticationException:
            print('服务器ip 或者 密码错误！')
            file_log('服务器ip 或者 密码错误！')
            return 1001
        except NoValidConnectionsError:
            print('服务器连接超时！')
            file_log('服务器连接超时！')
            return 1002
        except Exception as e:
            print(gl_name + " : " + host_ip + "其他错误 :", str(e))
            file_log("其他错误")
            return 1003
        return 1000

    # 这里是获取返回的CUP使用率、内存使用率、根目录百分比、data百分比
    def execute_some_command(self, command):
        stdin, stdout, stderr = self.ssh_client.exec_command(command)
        i = 0
        global gl_null, gl_ls, gl_cpu, gl_nc, gl_data, gl_data1
        gl_ls = ''

        for sysOutput in stdout.readlines():
            if sysOutput.startswith('[sscard]') or sysOutput.startswith(
                    'Command Startup Success') or sysOutput.startswith('-----'):
                continue
            if i == 0:
                gl_cpu = sysOutput.strip()
            elif i == 1:
                gl_nc = sysOutput.strip()
            elif i == 2:
                gl_null = sysOutput.strip()
            elif i == 3:
                gl_null = sysOutput.strip()
            elif i == 4:
                gl_data = sysOutput.strip() + '\n'
            elif i == 5:
                gl_data1 = sysOutput.strip() + '\n'
            else:
                gl_ls += sysOutput.strip()
            i += 1

    def ssh_logout(self):
        self.ssh_client.close()


class TestMysql:
    # 初始化变量
    def __init__(self, username, host, passwd, database):
        self.username = username
        self.host = host
        self.passwd = passwd
        self.database = database

    # 创建数据库连接
    def conn_mysql(self):  # 创建
        conn = pymysql.connect(user=self.username, host=self.host, password=self.passwd, db=self.database)
        return conn

    # 关闭数据库的提示信息
    def close_mysql(self):  # 关闭
        print("SQL Closed")

    def conn_mysql_1(self):  # 连接
        self.conn = self.conn_mysql()
        global cur
        cur = self.conn.cursor()

    # 查询数据
    def get_data_num(self):
        global cur, sql

        sql = 'select count(*) from text '
        cur.execute(sql)
        results = cur.fetchall()
        for i in results:
            global gl_f_id
            gl_f_id = i[0]
            print('服务器总数 【%s】' % gl_f_id)
            file_log('服务器总数 【%s】\n' % gl_f_id)

    # 这条我没有使用，我给删除调用了
    def get_charru(self, quyu_1, name, ip, user, passwd, i_prot, i_time):
        global cur

        # sql = "INSERT INTO f_handel VALUES ('" + quyu_1 + "','" + name + "','" + ip + "','" + user + "','" + passwd + "','ST64','" + i_prot + "');"
        sql = "INSERT INTO `ov`.`f_handel` ( `f_quyu`, `f_name`, `f_ip`, `f_user`, `f_passwd`, `f_status`, `f_port`, `f_time`) VALUES ( '" + quyu_1 + "', '" + name + "', '" + ip + "', '" + user + "', '" + passwd + "', 'ST64', '" + i_prot + "', '" + i_time + "');"
        sql1 = "commit;"
        # print(sql)
        try:
            cur.execute(sql)
            cur.execute(sql1)
            # print('数据插入成功')
            # file_log('数据插入成功\n')
        except:
            print('数据插入失败')
            file_log('数据插入失败\n')

    def get_data(self):  # 指定查询
        try:
            global cur, sql
            cur.execute(sql)
            results = cur.fetchall()
            for i in results:
                # print(str(i))
                global gl_id
                gl_id = i[0]
                global gl_name
                gl_name = i[1]
                global gl_ip
                gl_ip = i[2]
                global gl_user
                gl_user = i[3]
                global gl_passwd
                gl_passwd = i[4]
                global gl_port1
                gl_port1 = i[5]
                global gl_port2
                gl_port2 = i[6]
                global gl_port3
                gl_port3 = i[7]
                global f_path_1
                f_path_1 = i[8]
                global f_path_2
                f_path_2 = i[9]
                global f_path_3
                f_path_3 = i[10]
                global f_path_4
                f_path_4 = i[11]
                global f_path_5
                f_path_5 = i[12]
                global remote_ip
                remote_ip = i[13]
                global remote_username
                remote_username = i[14]
                global remote_password
                remote_password = i[15]
                global remote_backup_dir
                remote_backup_dir = i[16]
                global f_qy
                f_qy = i[17]
        except:
            print('')


# ping网络是否通，但是我改了返回值，因为可能禁ping
def m_wangluo(ip):  # 网络联通状态
    # message = ping(str(ip))
    # success_ping = "Reply"
    # if success_ping in str(message):
    #     print('服务器网络正常！(ping)')
    #     file_log('服务器网络正常！(ping)\n')
    #     return True
    # else:
    #     print('服务器网络异常！(ping)')
    #     file_log('服务器网络异常！(ping)\n')
    #     return True
    return True


# 判断端口是否正常
# def m_duankou_pd(ip, port):  # 判断端口是否开启
#     sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
#     result = sock.connect_ex((ip, port))
#     if result == 0:
#         # print(ip+': %s ： 正常'%port)
#         return 1
#     else:
#         # print(ip+': %s ： 异常'%port)
#         return 0


# 写log日志
def file_log(str):
    global stime1
    file_anme_log = open(stime1 + '.log', mode='a', encoding='UTF-8')

    file_anme_log.write(str)
    file_anme_log.close()


# 以备份文件名判断是否有今天或者昨天的文件   逻辑备份、物理备份、远程物理备份都走这个方法
def get_modification(today_files, one, ssh, wlibackup_wli, wlibackup_name):
    beifname = ''
    wlibackup_status = ''
    target_date = datetime.date.today()
    target_data3 = target_date.strftime("%Y%m%d")
    target_date1 = datetime.date.today() - datetime.timedelta(days=1)
    target_date4 = target_date1.strftime("%Y%m%d")

    if (one == "1"):
        beifname = "逻辑"
    if (one == "2"):
        beifname = "物理"
    if (one == "3"):
        beifname = "远程物理"

    files_modified_yesterday = []

    # 在远程服务器上执行命令以列出目标目录中的所有文件
    command = f"ls {today_files}"
    try:
        # 使用 SFTP 客户端
        # sftp = ssh.open_sftp()
        # 将命令发送到远程服务器   execute_some_command
        stdin, stdout, stderr = ssh.exec_command(command)
        file_list = stdout.read().decode().splitlines()
        today_ztai = ''
    except Exception as e:
        wlibackup_status = ''
        today_ztai = 'no'
        print(f"备份路径不存在： {e}")
        file_log('%s:的备份路径不存在' % (today_files) + '\n')
        return wlibackup_status, beifname

    if today_ztai == 'no':
        print(f"备份路径不存在", today_files)
        return wlibackup_status, beifname
    else:
        print(target_data3)
        print(target_date4)
        for filename in file_list:
            print(filename)
            if target_data3 in filename or target_date4 in filename:
                files_modified_yesterday = beifname + "有备份"
                if beifname == '物理':
                    wlibackup_wli = os.path.join(today_files, filename)
                    wlibackup_name = filename

        if not files_modified_yesterday:
            wlibackup_status = ''
            print(today_files, "下的", beifname, "备份异常.")
            file_log('%s备份异常' % (beifname) + '\n')
            beif = '%s备份异常' % beifname
        else:
            wlibackup_status = '物理备份正常'
            print("文件：", files_modified_yesterday)
            file_log('%s备份正常' % (beifname) + '\n')
            beif = '%s备份正常' % beifname

        print(wlibackup_wli)
        # 关闭 SFTP 客户端和 SSH 会话
        # sftp.close()

        return wlibackup_status, beif, wlibackup_wli, wlibackup_name


# 把结果写入excel
def pd_toexcel(data, filename):
    # 定义列名
    column_names = ['服务器名称', '服务器 IP', '内存占用率', 'CUP占用率', '根目录占用', 'Data占用', '逻辑备份',
                    '物理备份', '远程物理备份', '远程物理备份传输情况']

    # 检查文件路径是否包含扩展名
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'

    # 读取现有 Excel 文件
    if os.path.exists(filename):
        print("有文件")
        wb = load_workbook(filename)
        sheet = wb.active
    else:
        # 如果没有名为filename的Excel文件，创建一个新的文件
        print("没文件")
        wb = pd.DataFrame()
        wb.to_excel(filename)
        wb = load_workbook(filename)
        sheet = wb.active

    # 获取现有数据行数
    last_row = sheet.max_row

    # 创建数据帧
    df = pd.DataFrame([data], columns=[column_names])

    # 将数据帧追加到 Excel 文件
    for i, row in df.iterrows():
        sheet.append(row.tolist())

    # 保存更改
    wb.save(filename)


# 当远程备份没有时，进行远程备份操作
def rsync_transfer(wlibackup_wli):
    global wlibackup_wli1
    # 连接到服务器1
    ssh_server1 = paramiko.SSHClient()
    ssh_server1.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh_server1.connect(hostname=gl_ip, port=22, username=gl_user, password=gl_passwd, timeout=10)

    # 创建一个交互式的shell，输入执行连接第二台服务器的命令和密码
    transport = ssh_server1.get_transport()
    session = transport.open_session()
    session.set_combine_stderr(True)
    session.get_pty()
    session.invoke_shell()

    # 执行SFTP命令
    command = f'scp -v {wlibackup_wli} {remote_username}@{remote_ip}:{remote_today_files}'
    session.send(command + '\n')
    time.sleep(10)  # 随着网络环境的不同，这里可能需要适当调整

    # 读取从服务器上返回的消息
    data = session.recv(50000)
    detect = chardet.detect(data)
    msg = data.decode(detect['encoding'])
    if "Are you sure you want to continue connecting" in msg:
        session.send('yes\n')
        time.sleep(5)  # 这里可能需要适当调整时间

    print(msg)
    print("通过Scp传输到第二台服务器: ", remote_username)
    file_log("通过Scp传输到第二台服务器: %s" % remote_username + '\n')

    # 在 password 提示出现后，输入密码
    session.send(remote_password + '\n')
    time.sleep(2)
    print("输入了密码")
    file_log("输入了密码" + '\n')

    # 记录开始时间
    start_time = datetime.datetime.now()
    timeout = datetime.timedelta(minutes=30)

    while True:
        try:
            # Set a timeout
            session.settimeout(60)  # 设置1分钟的超时时间
            # 在发送文件后获取返回信息
            data = session.recv(50000)
            if not data:
                print("未收到任何数据，可能文件传输中断")
                file_log('未收到任何数据，可能文件传输中断' + '\n')
                wlibackup_wli1 = '未收到任何数据，可能文件传输中断'
                break
            detect = chardet.detect(data)
            try:
                returned_output = data.decode(detect['encoding'], errors='ignore')
            except UnicodeDecodeError:
                try:
                    returned_output = data.decode('GB18030', errors='ignore')
                except UnicodeDecodeError:
                    returned_output = data.decode('ISO-8859-1', errors='ignore')

            print(returned_output)

            # 检查返回信息
            if "debug1: Exit status 0" in returned_output:
                print("文件已成功传输")
                file_log('远程物理备份传输完成' + '\n')
                wlibackup_wli1 = '远程物理备份传输完成'
                break
            elif "debug1: Exit status 1" in returned_output:
                print("文件传输失败")
                file_log('远程物理备份传输失败' + '\n')
                wlibackup_wli1 = '远程物理备份传输失败'
                break

        except socket.timeout:
            print("接收数据操作已超时，停止等待")
            file_log('接收数据操作已超时，停止等待' + '\n')
            wlibackup_wli1 = '接收数据操作已超时，停止等待'
            break

        except Exception as e:
            print(f"发生错误: {str(e)}")
            file_log('SCP操作超时或出现其他错误' + '\n')
            wlibackup_wli1 = 'SCP操作超时或出现其他错误'
            break

        finally:
            # 检查是否超时
            if datetime.datetime.now() - start_time > timeout:
                print("已经超过半小时, 停止等待")
                file_log('已经超过半小时, 停止等待' + '\n')
                wlibackup_wli1 = '已经超过半小时, 停止等待'
                break

        # 延迟一段时间再进行下一次接收，避免过于密集的接收对 CPU 的影响
        time.sleep(30)

    # 关闭连接
    ssh_server1.close()
    return (wlibackup_wli1)


def connect_to_host(host_address, username, password):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        ssh.connect(hostname=host_address, username=username, password=password)
    except paramiko.AuthenticationException:
        file_log("Failed to connect to %s due to wrong username/password" % host_address + '\n')
        return None
    except paramiko.ssh_exception.NoValidConnectionsError:
        file_log(f"Failed to connect to {host_address}" + '\n')
        return None
    except paramiko.SSHException as e:
        file_log("Failed to connect to %s" % host_address + '\n')
        return None

    print("成功连接到跳板服务器: ", host_address)
    return ssh


def execute_some_command(session, command):
    session.send(command + "\n")
    time.sleep(1)  # adjust this timing as necessary
    data = session.recv(50000)
    detect = chardet.detect(data)
    output = data.decode(detect['encoding'])
    return output


def create_interactive_shell(ssh, server_info):
    transport = ssh.get_transport()
    session = transport.open_session()
    session.set_combine_stderr(True)
    session.get_pty()
    session.invoke_shell()

    # 在 terminal 上输入执行连接第二台服务器的命令
    session.send('ssh {}@{}\n'.format(server_info["username"], server_info["host_address"]))
    time.sleep(2)  # 随着网络环境的不同，这里可能需要适当调整

    # 读取从服务器上返回的消息
    msg = session.recv(1000).decode('utf-8')
    if "Are you sure you want to continue connecting" in msg:
        session.send('yes\n')
        time.sleep(1)  # 这里可能需要适当调整时间

    print("通过SSH连接到第二台服务器: ", server_info['host_address'])
    file_log("通过SSH连接到第二台服务器: %s" % (server_info['host_address']) + '\n')

    # 在 password 提示出现后，输入密码
    session.send(server_info["password"] + '\n')
    time.sleep(2)
    print("输入了密码")
    file_log("输入了密码" + '\n')

    return session


def main(servers):
    cpu = ''
    nc = ''
    genmul = ''
    guaz = ''
    luojibf = ''
    wulibf = ''

    # connect to the first server
    first_server = servers[0]
    ssh = connect_to_host(first_server["host_address"], first_server["username"], first_server["password"])
    if ssh == None:
        return cpu, nc, genmul, guaz, luojibf, wulibf
    else:
        second_server = servers[1]

        # 创建一个交互式的shell，输入执行连接第二台服务器的命令和密码
        session = create_interactive_shell(ssh, second_server)

        # 在连接的第二台服务器上运行命令
        command = f_path_1
        session.send(command + '\n')
        print("在第二台服务器上执行了命令: ", command)
        file_log("在第二台服务器上执行了命令: %s" % (command) + '\n')
        command = f_path_2
        session.send(command + '\n')
        time.sleep(1)
        print("在第二台服务器上执行了命令: ", command)
        file_log("在第二台服务器上执行了命令: %s" % (command) + '\n')
        time.sleep(1)
        # print the command result
        data = session.recv(10000)
        detect = chardet.detect(data)
        try:
            result = data.decode(detect['encoding'],errors='ignore')
        except UnicodeDecodeError:
            try:
                result = data.decode('UTF-8',errors='ignore')
            except UnicodeDecodeError:
                result = data.decode('ISO-8859-1',errors='ignore')  # 或者尝试'latin1'
        lines = result.splitlines()  # 使用 splitlines() 来分割结果
        if lines[-1] == '':  # 如果最后一行是空行
            lines = lines[:-1]  # 则去掉最后一行
        line_count = len(lines)  # 计算行数
        print(result)
        file_log(result + '\n')
        print(f"返回了 {line_count} 行")

        # 创建一个空列表用于存储匹配的值
        matched_values = []
        time.sleep(0.5)
        for line in lines:
            if re.match(r'^\d+(\.\d*)?%?$', line):  # 如果这一行以数字开头，可能包含一个小数点和小数，以%结尾
                print(line)  # 则输出这一行
                # 将匹配的值添加到列表中
                matched_values.append(line)
        # 如果列表的长度小于4，添加足够多的0.0
        while len(matched_values) < 4:
            matched_values.append("0.0")
        # 打印存储了所有匹配的行的列表
        print(matched_values)
        # file_log(matched_values + '\n')
        cpu = matched_values[0]
        nc = matched_values[1]
        genmul = matched_values[2]
        guaz = matched_values[3]

        # 然后执行 ls 命令
        command = 'ls ' + f_path_4
        session.send(command + '\n')
        time.sleep(1)
        data = session.recv(10000000)
        detect = chardet.detect(data)
        try:
            result1 = data.decode(detect['encoding'],errors='ignore')
        except UnicodeDecodeError:
            try:
                result1 = data.decode('GB18030',errors='ignore')
            except UnicodeDecodeError:
                result1 = data.decode('ISO-8859-1',errors='ignore')  # 或者尝试'latin1'

        # 后面的处理与之前一样
        lines1 = result1.splitlines()  # 使用 splitlines() 来分割结果
        if lines1[-1] == '':  # 如果最后一行是空行
            lines1 = lines1[:-1]  # 则去掉最后一行
        line_count1 = len(lines1)  # 计算行数
        # print(result1)
        # print(lines1)
        print(f"返回了 {line_count1} 行")

        # 移除第一行和最后一行
        data = lines1[1:-1]

        # 所有的文件名
        all_files = []

        for line in data:
            # 使用 split() 函数分割字符串
            files = line.split()

            # 将文件名添加至列表
            all_files.extend(files)
        time.sleep(1)
        print(all_files)
        # file_log("逻辑备份文件 %s" % all_files + '\n')
        files_modified_yesterday = pdwenjm(all_files)
        if files_modified_yesterday == '有备份':
            luojibf = "逻辑备份正常"
        else:
            luojibf = "逻辑备份异常"

        # 打印存储了所有匹配的行的列表
        # print(result1)

        # 然后执行 ls 命令
        command = 'ls ' + f_path_5
        session.send(command + '\n')
        time.sleep(1)
        data = session.recv(10000000)
        detect = chardet.detect(data)
        try:
            result1 = data.decode(detect['encoding'],errors='ignore')
        except UnicodeDecodeError:
            try:
                result1 = data.decode('GB18030',errors='ignore')
            except UnicodeDecodeError:
                result1 = data.decode('ISO-8859-1',errors='ignore')  # 或者尝试'latin1'

        # 后面的处理与之前一样
        lines1 = result1.splitlines()  # 使用 splitlines() 来分割结果
        if lines1[-1] == '':  # 如果最后一行是空行
            lines1 = lines1[:-1]  # 则去掉最后一行
        line_count1 = len(lines1)  # 计算行数
        # print(result1)
        # print(lines1)
        print(f"返回了 {line_count1} 行")

        # 移除第一行和最后一行
        data = lines1[1:-1]

        # 所有的文件名
        all_files = []

        for line in data:
            # 使用 split() 函数分割字符串
            files = line.split()

            # 将文件名添加至列表
            all_files.extend(files)

        print(all_files)
        # file_log("物理备份文件 %s" % all_files + '\n')
        files_modified_yesterday = pdwenjm(all_files)
        if files_modified_yesterday == '有备份':
            wulibf = "物理备份正常"
        else:
            wulibf = "物理备份异常"

        return cpu, nc, genmul, guaz, luojibf, wulibf




def pdwenjm(file_list):
    target_data3 = target_date.strftime("%Y%m%d")
    target_date1 = datetime.date.today() - datetime.timedelta(days=1)
    target_date4 = target_date1.strftime("%Y%m%d")
    for filename in file_list:
        print(filename)
        if target_data3 in filename or target_date4 in filename:
            files_modified_yesterday = "有备份"
            return files_modified_yesterday


if __name__ == "__main__":
    xxx = input('请输入循环次数 ：\n')
    stime1 = time.strftime("%Y-%m-%d_%H_%M_%S", time.localtime(time.time()))
    username = 'root'
    host = 'localhost'
    passwd = '123456'
    database = 'ov'
    try:
        mysql = TestMysql(username, host, passwd, database)
        mysql.conn_mysql()
        mysql.conn_mysql_1()
        mysql.get_data_num()
    except pymysql.err.ProgrammingError as e:
        print("Exception Error is %s" % (e))
        file_log("Exception Error is %s" % (e))
    except pymysql.err.OperationalError as e:
        print("Exception Error is %s" % (e))
        file_log("Exception Error is %s" % (e))
    for i in range(0, gl_f_id, 1):  # 查询服务几台
        # sql = 'select * from text where id = %s' % i
        sql = 'select id,f_name,f_ip,f_user,f_passwd,f_port1,f_port2,f_port3,f_path_1,f_path_2,f_path_3,f_path_4,f_path_5,remote_ip, remote_username, remote_password, remote_backup_dir,f_quyu from text where id = %s ' % i
        try:
            mysql.get_data()
        except pymysql.err.ProgrammingError as e:
            print("Exception Error is %s" % (e))
            file_log("Exception Error is %s\n" % (e))
        except pymysql.err.OperationalError as e:
            print("Exception Error is %s" % (e))
            file_log("Exception Error is %s\n" % (e))
        # 判断网络是否正常
        print('正在检查服务器：%s ip:%s' % (gl_name, gl_ip))
        file_log('正在检查服务器：%s ip:%s\n' % (gl_name, gl_ip))
        fwq_name = gl_name
        if m_wangluo(gl_ip):
            # 判断端口是否正常
            # mysql.get_charru('贵阳', 'root', '123456', '123456', '11111', '9999')
            stime = time.strftime("%Y/%m/%d %H:%M:%S", time.localtime(time.time()))
            # if gl_port1 != 0:
            #     if m_duankou_pd(gl_ip, gl_port1) == 1:
            #         gg_port_1 = '端口正常'
            #     else:
            #         gg_port_1 = '端口异常'
            # print(f_qy, gl_name, gl_ip, gl_user, gl_passwd, gl_port1)
            # mysql.get_charru(f_qy, gl_name, gl_ip, gl_user, gl_passwd, str(gl_port1),stime)
            # if gl_port2 != 0:
            #     if m_duankou_pd(gl_ip, gl_port2) == 1:
            #         gg_port_2 = '端口正常'
            #     else:
            #         gg_port_2 = '端口异常'
            # mysql.get_charru(f_qy, gl_name, gl_ip, gl_user, gl_passwd, str(gl_port2),stime)
            # mysql.get_charru(f_qy, gl_name, gl_ip, gl_user, gl_passwd, gl_port2)
            # if gl_port3 != 0:
            #     if m_duankou_pd(gl_ip, gl_port3) == 1:
            #         gg_port_3 = '端口正常'
            #     else:
            #         gg_port_3 = '端口异常'
            # mysql.get_charru(f_qy, gl_name, gl_ip, gl_user, gl_passwd, gl_port3)
            # mysql.get_charru(f_qy, gl_name, gl_ip, gl_user, gl_passwd, str(gl_port3),stime)
            ssh = SshClient()

            print('巡检时间：%s\nip:%s\n%s:%s\n%s:%s\n%s:%s' % (
                stime, gl_ip, gl_port1, gg_port_1, gl_port2, gg_port_2, gl_port3, gg_port_3))
            file_log('巡检时间：%s\nip:%s\n%s:%s\n%s:%s\n%s:%s\n' % (
                stime, gl_ip, gl_port1, gg_port_1, gl_port2, gg_port_2, gl_port3, gg_port_3))
            gl_name = ''
            command = '%s%s%s' % (f_path_1, f_path_2, f_path_3,)
            # print('command : %s' % command)
            if ssh.ssh_login(gl_ip, gl_user, gl_passwd) == 1000:
                ssh.execute_some_command(command)
                # ssh.ssh_logout()
                print('CPU使用率：%s\n内存使用率: %s\n根目录率: %s\n挂载盘使用率: %s' % (
                    gl_cpu, gl_nc, gl_data, gl_data1))
                file_log('CPU使用率：%s\n内存使用率: %s\n根目录率: %s\n挂载盘使用率: %s\n' % (
                    gl_cpu, gl_nc, gl_data, gl_data1))
                print('服务器备份情况：\n%s' % gl_ls)
                file_log('服务器备份情况：\n%s\n' % gl_ls)

                wlibackup_wli = ''
                wlibackup_name = ''
                wlibackup_status = ''
                if f_path_4 != '':
                    today_files = f'{f_path_4}/'
                    wlibackup_status, beif1, wlibackup_wli, wlibackup_name = get_modification(today_files, "1",
                                                                                              ssh.ssh_client,
                                                                                              wlibackup_wli,
                                                                                              wlibackup_name)
                    print('服务器备份路径不为空')

                if f_path_5 != '':
                    today_files = f'{f_path_5}/'
                    wlibackup_status, beif2, wlibackup_wli, wlibackup_name = get_modification(today_files, "2",
                                                                                              ssh.ssh_client,
                                                                                              wlibackup_wli,
                                                                                              wlibackup_name)
                    print('服务器备份路径不为空')

                # 如果本地备份正常，检查远程服务器备份
                if wlibackup_status == '物理备份正常':
                    try:
                        ssh_remote = paramiko.SSHClient()
                        ssh_remote.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        ssh_remote.connect(remote_ip, username=remote_username, password=remote_password, timeout=15)
                        remote_zhuangtai = ''
                    except Exception as e:
                        remote_zhuangtai = 'no'
                        print(f"远程服务器连接失败: {e}")

                    if remote_zhuangtai == 'no':
                        print(f"远程服务器连接失败")
                        file_log('远程服务器连接失败\n%s\n')
                    else:
                        # 构造目标目录路径
                        remote_today_files = f'{remote_backup_dir}/'
                        if remote_backup_dir != '':
                            wlibackup_status, beif3, wlibackup_wli, wlibackup_name = get_modification(
                                remote_today_files, "3", ssh_remote, wlibackup_wli, wlibackup_name)
                            print('远程服务器备份路径不为空')
                        # 关闭会话和连接
                        ssh_remote.close()

                else:
                    remote_backup_status = '未检查远程备份'
                    print("服务器远程检查情况:", remote_backup_status)
                    file_log('服务器远程检查情况：\n%s' % remote_backup_status)

                if beif3 == '远程物理备份异常':
                    wlibackup_wli = rsync_transfer(wlibackup_wli)
                    # wlibackup_wli = '远程物理备份已传输完成'

                # print(filename)
                testData = [fwq_name, gl_ip, gl_nc, gl_cpu, gl_data, gl_data1, beif1, beif2, beif3, wlibackup_wli]
                # print(testData)
                pd_toexcel(testData, filename + '.xlsx')
                gl_cpu = ''
                gl_nc = ''
                gl_data = ''
                gl_data1 = ''
                gl_ls = ''
                gl_null = ''
                beif1 = ''
                beif2 = ''
                beif3 = ''
                wlibackup_wli = ''
            else:
                servers = [{"host_address": remote_ip, "username": remote_username, "password": remote_password},
                           {"host_address": gl_ip, "username": gl_user, "password": gl_passwd}]
                gl_cpu, gl_nc, gl_data, gl_data1, beif1, beif2 = main(servers)
                testData = [fwq_name, gl_ip, gl_nc, gl_cpu, gl_data, gl_data1, beif1, beif2, beif3, wlibackup_wli]
                pd_toexcel(testData, filename + '.xlsx')
                gl_cpu = ''
                gl_nc = ''
                gl_data = ''
                gl_data1 = ''
                gl_ls = ''
                gl_null = ''
                beif1 = ''
                beif2 = ''
                beif3 = ''
                wlibackup_wli = ''

                # continue




        else:
            print('%s：服务器网络异常 ip:%s' % (gl_name, gl_ip))
            file_log('%s：服务器网络异常 ip:%s\n' % (gl_name, gl_ip))
        print('%s 服务器检查结束!' % gl_name)
        file_log('%s 服务器检查结束!\n' % gl_name)
        print('\r\n------------------------------------------------------\r\n')
        file_log('\r\n------------------------------------------------------\r\n')
        # print(gl_ip + ':%s'%gl_port2)
    # mysql.close_mysql()
    # print('循环 【%s】 次  已完成\n'%xxx)
    # file_log('循环 【%s】 次  已完成\n'%xxx)
    input('按任意键退出：')
